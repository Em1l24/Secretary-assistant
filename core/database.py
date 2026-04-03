# core/database.py
import openpyxl
import os
import sys
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Any

class Database:
    """Класс для работы с Excel-базой данных"""
    
    def __init__(self, db_path: str = None):
        if db_path is None:
            if getattr(sys, 'frozen', False):
                data_dir = Path.home() / "Documents" / "GECAssistant"
            else:
                if hasattr(sys, '_MEIPASS'):
                    data_dir = Path(sys._MEIPASS).parent / "data"
                else:
                    data_dir = Path(__file__).parent.parent / "data"
            try:
                data_dir.mkdir(parents=True, exist_ok=True)
            except PermissionError:
                import tempfile
                data_dir = Path(tempfile.gettempdir()) / "GECAssistant"
                data_dir.mkdir(exist_ok=True)
            db_path = str(data_dir / "БД.xlsx")
        
        self.db_path = db_path
        self.first_run_flag = str(Path(db_path).parent / ".first_run")
        self._check_first_run()
        self._ensure_db_exists()
    
    def _check_first_run(self):
        try:
            if not os.path.exists(self.first_run_flag):
                print("🔄 ПЕРВЫЙ ЗАПУСК - очищаем все данные...")
                if os.path.exists(self.db_path):
                    self._clear_all_data()
                Path(self.first_run_flag).parent.mkdir(parents=True, exist_ok=True)
                with open(self.first_run_flag, 'w', encoding='utf-8') as f:
                    f.write('initialized')
                print("✅ Флаг первого запуска создан")
            else:
                print("📁 Флаг найден - данные не очищаем")
        except Exception as e:
            print(f"⚠️ Ошибка проверки первого запуска: {e}")
    
    def _clear_all_data(self):
        try:
            wb = self._load_workbook()
            for sheet_name in ['тест', 'экзамен', 'ВКР']:
                if sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    max_row = sheet.max_row
                    if max_row >= 4:
                        for row in range(max_row, 3, -1):
                            sheet.delete_rows(row)
            if 'комиссия' in wb.sheetnames:
                sheet = wb['комиссия']
                max_row = sheet.max_row
                if max_row > 3:
                    for row in range(max_row, 3, -1):
                        sheet.delete_rows(row)
            wb.save(self.db_path)
            print("✅ Все данные очищены")
            return True
        except Exception as e:
            print(f"❌ Ошибка очистки данных: {e}")
            return False
    
    def _ensure_db_exists(self):
        try:
            db_dir = Path(self.db_path).parent
            db_dir.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            print(f"⚠️ Не удалось создать папку: {e}")
        
        if not os.path.exists(self.db_path):
            print("📁 Создаём новый Excel файл...")
            wb = openpyxl.Workbook()
            for sheet_name in ['тест', 'экзамен', 'ВКР', 'комиссия']:
                if sheet_name not in wb.sheetnames:
                    wb.create_sheet(sheet_name)
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
            self._create_headers(wb)
            wb.save(self.db_path)
            print("✅ Excel файл создан")
    
    def _create_headers(self, wb):
        vkr_headers = ['ФИО', 'Протокол №', 'Направление подготовки', 'Тема дипломной работы', 'Научный руководитель', 'Дата защиты', 'Состав ГЭК утвержден приказом от', 'Должность, место работы дипломного руководителя', 'При консультации', 'Обучающийся допущен до защиты ВКР приказом от', 'Кол-во страниц', 'Чертежи', 'Иллюстрационный материал', 'Отзыв руководителя', 'Время сообщения ВКР', 'Заданные вопросы', 'Характеристика ответов обучающегося', 'Оценка', 'Квалификация', 'Выдать диплом', 'Отметить, что']
        test_headers = ['ФИО', 'Протокол №', 'Направление подготовки', 'Группа', 'Дата экзамена', 'Состав ГЭК утвержден приказом от', 'Оценка', 'Характеристика ответов обучающегося']
        exam_headers = ['ФИО', 'Протокол №', 'Направление подготовки', 'Группа', 'Дата экзамена', 'Состав ГЭК утвержден приказом от', '№ вытянутого билета', 'Дата утверждения билетов', 'Вопросы билета', 'Дополнительные вопросы', 'Характеристика ответов обучающегося', 'Оценка']
        commission_headers = ['ГЭК ТИУ', 'ГЭК ТИУ']
        self._write_headers(wb, 'ВКР', vkr_headers, start_row=4)
        self._write_headers(wb, 'тест', test_headers, start_row=4)
        self._write_headers(wb, 'экзамен', exam_headers, start_row=4)
        self._write_headers(wb, 'комиссия', commission_headers, start_row=1)
    
    def _write_headers(self, wb, sheet_name: str, headers: List[str], start_row: int = 1):
        sheet = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
        for col, header in enumerate(headers, 1):
            sheet.cell(row=start_row, column=col, value=header)
            sheet.cell(row=start_row, column=col).font = openpyxl.styles.Font(bold=True)
            sheet.cell(row=start_row, column=col).alignment = openpyxl.styles.Alignment(wrap_text=True)
    
    def _load_workbook(self) -> openpyxl.Workbook:
        return openpyxl.load_workbook(self.db_path)
    
    def _save_workbook(self, wb: openpyxl.Workbook):
        wb.save(self.db_path)
    
    def save_commission(self, commission_data: Dict[str, Any]):
        wb = self._load_workbook()
        sheet = wb['комиссия']
        for row in range(3, 21):
            for col in [1, 2]:
                sheet.cell(row=row, column=col).value = None
        if commission_data.get('chairman'): sheet['B3'].value = commission_data['chairman']
        if commission_data.get('chairman_position'): sheet['B4'].value = commission_data['chairman_position']
        if commission_data.get('secretary'): sheet['B16'].value = commission_data['secretary']
        if commission_data.get('secretary_position'): sheet['B17'].value = commission_data['secretary_position']
        member_mapping = {1: (6, 7), 2: (9, 10), 3: (12, 13), 4: (19, 20)}
        for i in range(1, 5):
            name_key = f'member_{i}'
            pos_key = f'member_{i}_position'
            name_row, pos_row = member_mapping[i]
            name_val = commission_data.get(name_key)
            pos_val = commission_data.get(pos_key)
            if name_val is not None: sheet[f'B{name_row}'].value = str(name_val).strip()
            if pos_val is not None: sheet[f'B{pos_row}'].value = str(pos_val).strip()
        self._save_workbook(wb)
    
    def get_commission(self) -> Dict[str, Any]:
        wb = self._load_workbook()
        sheet = wb['комиссия']
        data = {
            'chairman': sheet['B3'].value or '',
            'chairman_position': sheet['B4'].value or '',
            'secretary': sheet['B16'].value or '',
            'secretary_position': sheet['B17'].value or '',
        }
        member_mapping = {1: (6, 7), 2: (9, 10), 3: (12, 13), 4: (19, 20)}
        for i in range(1, 5):
            name_row, pos_row = member_mapping[i]
            name_val = sheet[f'B{name_row}'].value
            pos_val = sheet[f'B{pos_row}'].value
            data[f'member_{i}'] = name_val if name_val is not None else ''
            data[f'member_{i}_position'] = pos_val if pos_val is not None else ''
        return data
    
    def save_common_data_gos(self, common_data: Dict[str, Any], exam_state_date: str = None):
        wb = self._load_workbook()
        sheet = wb['комиссия']
        for row in range(21, 34):
            for col in [1, 2]:
                sheet.cell(row=row, column=col).value = None
        fields_mapping = {
            'direction': 'Направление подготовки',
            'group': 'Группа',
            'date': 'Дата экзамена',
            'dategek': 'Состав ГЭК утвержден приказом от'
        }
        for i, (key, label) in enumerate(fields_mapping.items()):
            row = 21 + i
            value = common_data.get(key, '')
            if value:
                sheet.cell(row=row, column=1, value=label)
                sheet.cell(row=row, column=2, value=str(value).strip())
        if exam_state_date:
            sheet.cell(row=26, column=1, value='Дата утверждения билетов')
            sheet.cell(row=26, column=2, value=str(exam_state_date).strip())
        self._save_workbook(wb)
    
    def get_common_data_gos(self) -> Dict[str, Any]:
        wb = self._load_workbook()
        sheet = wb['комиссия']
        data = {}
        fields = ['direction', 'group', 'date', 'dategek']
        for i, field in enumerate(fields):
            value = sheet.cell(row=21 + i, column=2).value
            data[field] = value if value is not None else ''
        state_date = sheet.cell(row=26, column=2).value
        data['state_date'] = state_date if state_date is not None else ''
        return data
    
    def save_common_data_vkr(self, common_data: Dict[str, Any]):
        wb = self._load_workbook()
        sheet = wb['комиссия']
        start_row = 28
        for row in range(start_row, start_row + 5):
            for col in [1, 2]:
                sheet.cell(row=row, column=col).value = None
        fields_mapping = {
            'direction': 'Направление подготовки',
            'date': 'Дата защиты',
            'dategek': 'Состав ГЭК утвержден приказом от',
            'order': 'Допущен до защиты приказом от',
            'quali': 'Квалификация'
        }
        for i, (key, label) in enumerate(fields_mapping.items()):
            row = start_row + i
            value = common_data.get(key, '')
            if value:
                sheet.cell(row=row, column=1, value=label)
                sheet.cell(row=row, column=2, value=str(value).strip())
        self._save_workbook(wb)
    
    def get_common_data_vkr(self) -> Dict[str, Any]:
        wb = self._load_workbook()
        sheet = wb['комиссия']
        start_row = 28
        fields = ['direction', 'date', 'dategek', 'order', 'quali']
        data = {}
        for i, field in enumerate(fields):
            value = sheet.cell(row=start_row + i, column=2).value
            data[field] = value if value is not None else ''
        return data
    
    def clear_all_common_data(self):
        try:
            wb = self._load_workbook()
            sheet = wb['комиссия']
            max_row = sheet.max_row
            if max_row > 2:
                for row in range(max_row, 2, -1):
                    for col in [1, 2]:
                        sheet.cell(row=row, column=col).value = None
            self._save_workbook(wb)
            print("✅ Все данные на листе 'комиссия' очищены")
            return True
        except Exception as e:
            print(f"❌ Ошибка очистки данных: {e}")
            return False
    
    def clear_common_data_only(self):
        try:
            wb = self._load_workbook()
            sheet = wb['комиссия']
            for row in range(21, 41):
                for col in [1, 2]:
                    sheet.cell(row=row, column=col).value = None
            self._save_workbook(wb)
            print("✅ Общие данные очищены")
            return True
        except Exception as e:
            print(f"❌ Ошибка очистки общих данных: {e}")
            return False
    
    def add_student(self, sheet_name: str, student_data: Dict[str, Any]) -> bool:
        try:
            wb = self._load_workbook()
            sheet = wb[sheet_name]
            last_row = self._get_last_row(sheet)
            new_row = last_row + 1
            student_data = self._fill_from_common_data(sheet_name, student_data)
            self._fill_student_row(sheet, sheet_name, new_row, student_data)
            self._save_workbook(wb)
            return True
        except Exception as e:
            print(f"❌ Ошибка добавления студента: {e}")
            return False
    
    def _fill_from_common_data(self, sheet_name: str, student_data: Dict[str, Any]) -> Dict[str, Any]:
        if sheet_name == 'тест':
            common = self.get_common_data_gos()
            for key in ['direction', 'group', 'date', 'dategek']:
                if key not in student_data or not student_data[key]:
                    student_data[key] = common.get(key, '')
        elif sheet_name == 'экзамен':
            common = self.get_common_data_gos()
            for key in ['direction', 'group', 'date', 'dategek', 'state_date']:
                if key not in student_data or not student_data[key]:
                    student_data[key] = common.get(key, '')
        elif sheet_name == 'ВКР':
            common = self.get_common_data_vkr()
            for key in ['direction', 'date', 'dategek', 'order', 'quali']:
                if key not in student_data or not student_data[key]:
                    student_data[key] = common.get(key, '')
        return student_data
    
    def _fill_student_row(self, sheet, sheet_name: str, row: int, student_data: Dict[str, Any]):
        if sheet_name == 'ВКР':
            mapping = [('fio', 1), ('protocol', 2), ('direction', 3), ('theme', 4), ('leader', 5), ('date', 6), ('dategek', 7), ('post', 8), ('con', 9), ('order', 10), ('pages', 11), ('con_1', 12), ('con_2', 13), ('review', 14), ('time', 15), ('questions', 16), ('property', 17), ('point', 18), ('quali', 19), ('diplom', 20), ('tom', 21)]
        elif sheet_name == 'тест':
            mapping = [('fio', 1), ('protocol', 2), ('direction', 3), ('group', 4), ('date', 5), ('dategek', 6), ('mark', 7), ('property', 8)]
        elif sheet_name == 'экзамен':
            mapping = [('fio', 1), ('protocol', 2), ('direction', 3), ('group', 4), ('date', 5), ('dategek', 6), ('ticket', 7), ('state_date', 8), ('questions', 9), ('add_questions', 10), ('property', 11), ('mark', 12)]
        else:
            return
        for key, col in mapping:
            value = student_data.get(key, '')
            sheet.cell(row=row, column=col, value=value)
    
    def get_all_students(self, sheet_name: str) -> List[Dict[str, Any]]:
        try:
            wb = self._load_workbook()
            sheet = wb[sheet_name]
            last_row = self._get_last_row(sheet)
            students = []
            for row in range(5, last_row + 1):
                fio = sheet.cell(row=row, column=1).value
                if fio is not None and str(fio).strip() != '' and str(fio).strip() != 'ФИО':
                    student = self._read_student_row(sheet, sheet_name, row)
                    student['_row'] = row
                    students.append(student)
            return students
        except Exception as e:
            return []
    
    def _read_student_row(self, sheet, sheet_name: str, row: int) -> Dict[str, Any]:
        if sheet_name == 'ВКР':
            return {'fio': sheet.cell(row=row, column=1).value, 'protocol': sheet.cell(row=row, column=2).value, 'direction': sheet.cell(row=row, column=3).value, 'theme': sheet.cell(row=row, column=4).value, 'leader': sheet.cell(row=row, column=5).value, 'date': sheet.cell(row=row, column=6).value, 'dategek': sheet.cell(row=row, column=7).value, 'post': sheet.cell(row=row, column=8).value, 'con': sheet.cell(row=row, column=9).value, 'order': sheet.cell(row=row, column=10).value, 'pages': sheet.cell(row=row, column=11).value, 'con_1': sheet.cell(row=row, column=12).value, 'con_2': sheet.cell(row=row, column=13).value, 'review': sheet.cell(row=row, column=14).value, 'time': sheet.cell(row=row, column=15).value, 'questions': sheet.cell(row=row, column=16).value, 'property': sheet.cell(row=row, column=17).value, 'point': sheet.cell(row=row, column=18).value, 'quali': sheet.cell(row=row, column=19).value, 'diplom': sheet.cell(row=row, column=20).value, 'tom': sheet.cell(row=row, column=21).value}
        elif sheet_name == 'тест':
            return {'fio': sheet.cell(row=row, column=1).value, 'protocol': sheet.cell(row=row, column=2).value, 'direction': sheet.cell(row=row, column=3).value, 'group': sheet.cell(row=row, column=4).value, 'date': sheet.cell(row=row, column=5).value, 'dategek': sheet.cell(row=row, column=6).value, 'mark': sheet.cell(row=row, column=7).value, 'property': sheet.cell(row=row, column=8).value}
        elif sheet_name == 'экзамен':
            return {'fio': sheet.cell(row=row, column=1).value, 'protocol': sheet.cell(row=row, column=2).value, 'direction': sheet.cell(row=row, column=3).value, 'group': sheet.cell(row=row, column=4).value, 'date': sheet.cell(row=row, column=5).value, 'dategek': sheet.cell(row=row, column=6).value, 'ticket': sheet.cell(row=row, column=7).value, 'state_date': sheet.cell(row=row, column=8).value, 'questions': sheet.cell(row=row, column=9).value, 'add_questions': sheet.cell(row=row, column=10).value, 'property': sheet.cell(row=row, column=11).value, 'mark': sheet.cell(row=row, column=12).value}
        return {}
    
    def delete_student(self, sheet_name: str, row_number: int) -> bool:
        try:
            wb = self._load_workbook()
            sheet = wb[sheet_name]
            if row_number < 5 or row_number > sheet.max_row: return False
            if sheet.cell(row=row_number, column=1).value is None: return False
            sheet.delete_rows(row_number, 1)
            self._save_workbook(wb)
            return True
        except: return False
    
    def update_student(self, sheet_name: str, row_number: int, student_data: Dict[str, Any]) -> bool:
        try:
            wb = self._load_workbook()
            sheet = wb[sheet_name]
            if row_number < 5 or row_number > sheet.max_row: return False
            student_data = self._fill_from_common_data(sheet_name, student_data)
            self._fill_student_row(sheet, sheet_name, row_number, student_data)
            self._save_workbook(wb)
            return True
        except: return False
    
    def clear_sheet(self, sheet_name: str) -> bool:
        try:
            wb = self._load_workbook()
            sheet = wb[sheet_name]
            last_row = self._get_last_row(sheet)
            if last_row >= 5: sheet.delete_rows(5, last_row - 4)
            self._save_workbook(wb)
            return True
        except: return False
    
    def _get_last_row(self, sheet, start_row: int = 4) -> int:
        max_row = sheet.max_row
        if max_row < start_row: return start_row - 1
        for row in range(max_row, start_row, -1):
            cell_value = sheet.cell(row=row, column=1).value
            if cell_value is not None and str(cell_value).strip() != '' and str(cell_value).strip() != 'ФИО':
                return row
        return start_row